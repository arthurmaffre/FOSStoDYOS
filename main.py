import streamlit as st
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import io
import base64
import os
import glob

# Import for auto-refresh
from streamlit_autorefresh import st_autorefresh

st.set_page_config(page_title="Export Moûts FOSS vers Dyostem", layout="wide", initial_sidebar_state="collapsed")

def create_file(ws, Nom_colonne):
    import string
    alphabet = list(string.ascii_uppercase)[:len(Nom_colonne)]
    for count, i in enumerate(Nom_colonne):
        ws[alphabet[count] + str(1)] = i
    # Set date format on header cell for the date column
    ws.cell(row=1, column=3).number_format = 'dd/mm/yyyy;@'
    return ws

def get_dates_from_file(file_content):
    df = pd.read_csv(io.BytesIO(file_content), sep=';', encoding='iso-8859-1')
    
    dates = set()
    for _, row in df.iterrows():
        if row['Product'] == "Moûts":
            date_value = row['Date']
            if date_value:
                try:
                    parsed_date = datetime.strptime(date_value, '%d/%m/%Y').strftime('%d/%m/%Y')
                    dates.add(parsed_date)
                except ValueError:
                    pass
    sorted_dates_str = sorted(list(dates))
    available_dates_obj = [datetime.strptime(d, '%d/%m/%Y').date() for d in sorted_dates_str]
    return sorted_dates_str, available_dates_obj

@st.cache_data
def process_file(file_content, selected_date):
    Nom_colonne = ["Nom de parcelle", "Cépage", "Date analyse", "Code échantillon", "Quantité Sucre (mg/baie)", "TAP (% vol)", "Acidité totale (g H2SO4/l)", "pH", "Acide malique (g/l)", "Acide tartrique", "Azote assimilable (mg/l)", "Potassium (g/l)", "Anthocyanes (mg/l)"]
    
    wb = Workbook()
    ws = wb.active
    ws = create_file(ws, Nom_colonne)
    
    # Pre-format the date column (column 3) as date for rows 2 to 1000
    for row_num in range(2, 1001):
        ws.cell(row=row_num, column=3).number_format = 'dd/mm/yyyy;@'
    
    df = pd.read_csv(io.BytesIO(file_content), sep=';', encoding='iso-8859-1')
    
    ligne_ws = 1
    for _, row in df.iterrows():
        if row['Product'] == "Moûts":
            date_value = row['Date']
            if date_value and datetime.strptime(date_value, '%d/%m/%Y').strftime('%d/%m/%Y') == selected_date:
                ligne_ws += 1
                ws.cell(row=ligne_ws, column=1).value = row['ID']  # Nom de parcelle
                ws.cell(row=ligne_ws, column=2).value = "Sauvignon blanc"  # Cépage: Hardcoded
                
                # Set date as datetime object for proper Excel date format
                date_obj = datetime.strptime(date_value, '%d/%m/%Y')
                ws.cell(row=ligne_ws, column=3).value = date_obj
                ws.cell(row=ligne_ws, column=3).number_format = 'dd/mm/yyyy;@'  # Set display format with ';@' to recognize as Date category
                
                ws.cell(row=ligne_ws, column=5).value = row['Sucre']  # Quantité Sucre
                ws.cell(row=ligne_ws, column=6).value = row['TAP']  # TAP
                ws.cell(row=ligne_ws, column=7).value = row['AT']  # Acidité totale
                ws.cell(row=ligne_ws, column=8).value = row['pH']  # pH
                ws.cell(row=ligne_ws, column=9).value = row['AMal']  # Acide malique
                ws.cell(row=ligne_ws, column=10).value = row['Tartaric']  # Acide tartrique
                ws.cell(row=ligne_ws, column=11).value = row['N_ass']  # Azote assimilable
                ws.cell(row=ligne_ws, column=12).value = row['K']  # Potassium
                ws.cell(row=ligne_ws, column=13).value = row['Anth']  # Anthocyanes
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def format_filename(name):
    if name.startswith("Moûts") and len(name) >= 20:  # Moûts + 14 digits + .csv
        ts = name[5:-4]  # Extract timestamp part: YYYYMMDDHHMMSS
        if len(ts) >= 14 and ts.isdigit():
            ts = ts[:14]  # Take only first 14 digits if longer
            year = ts[0:4]
            month = ts[4:6]
            day = int(ts[6:8])
            hour = int(ts[8:10])
            minute = int(ts[10:12])
            second = int(ts[12:14])
            months = {
                '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
                '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
                '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
            }
            month_str = months.get(month, month)
            date_str = f"{day} {month_str} {year}"
            time_str = f"{hour}h{minute}m{second}s"
            return f"Moûts - {date_str} - {time_str}"
    return name

# Custom CSS for Apple-like modern UI
st.markdown("""
    <style>
    .stApp {
        background-color: #f5f5f7;
        color: #1d1d1f;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
    }
    .stButton > button {
        background-color: #0071e3;
        color: white;
        border-radius: 8px;
        border: none;
        padding: 10px 24px;
        font-size: 16px;
        font-weight: 600;
        transition: background-color 0.3s;
    }
    .stButton > button:hover {
        background-color: #0077ed;
    }
    .stDateInput > div > div {
        background-color: white;
        border-radius: 8px;
        border: 1px solid #d2d2d7;
    }
    .stFileUploader > div > div {
        background-color: white;
        border-radius: 8px;
        border: 1px solid #d2d2d7;
        padding: 10px;
    }
    h1 {
        color: #1d1d1f;
        font-size: 32px;
        font-weight: 600;
        text-align: center;
        margin-bottom: 40px;
    }
    .stDownloadButton > button {
        background-color: #0071e3;
        color: white;
        border-radius: 8px;
        border: none;
        padding: 10px 24px;
        font-size: 16px;
        font-weight: 600;
        transition: background-color 0.3s;
        width: 100%;
    }
    .stDownloadButton > button:hover {
        background-color: #0077ed;
    }
    </style>
""", unsafe_allow_html=True)

st.title("Export Moûts FOSS vers Dyostem")

# Define the specific directory to scan
export_dir = r"C:\FOSS\FossIntegrator\Export"

# Scan the specific directory for CSV files starting with "Moûts", sorted by modification time descending (latest first)
if os.path.exists(export_dir):
    all_csv_files = glob.glob(os.path.join(export_dir, "*.csv"))
    mouts_files = [f for f in all_csv_files if os.path.basename(f).startswith("Moûts")]
    csv_files = sorted(mouts_files, key=os.path.getmtime, reverse=True)[:6]  # Limit to the 6 most recent
    file_basenames = [os.path.basename(f) for f in csv_files]
    displayed_options = [format_filename(b) for b in file_basenames]
else:
    csv_files = []
    file_basenames = []
    displayed_options = []
    st.error(f"Le répertoire {export_dir} n'existe pas. Veuillez vérifier le chemin.")

col1, col2 = st.columns(2)

with col1:
    selected_display = st.selectbox(
        "Sélectionner un fichier CSV du répertoire Export (dernier en premier)",
        options=displayed_options,
        index=0 if displayed_options else None,
        disabled=not displayed_options
    )

with col2:
    uploaded_file = st.file_uploader("Ou charger un fichier (.csv)", type="csv")

file_content = None
source = None
selected_file = None

if uploaded_file:
    file_content = uploaded_file.getvalue()
    source = "uploaded"
elif selected_display:
    index = displayed_options.index(selected_display)
    selected_file = file_basenames[index]
    file_path = os.path.join(export_dir, selected_file)
    with open(file_path, "rb") as f:
        file_content = f.read()
    source = "selected"

formatted_name = None
if file_content:
    if source == "selected":
        formatted_name = selected_display
    elif source == "uploaded":
        potential_formatted = format_filename(uploaded_file.name)
        if potential_formatted != uploaded_file.name:
            formatted_name = potential_formatted
    
    if source == "selected":
        st.info(f"Fichier sélectionné : {selected_display}")
    with st.spinner("Analyse du fichier..."):
        sorted_dates_str, available_dates_obj = get_dates_from_file(file_content)
    
    if available_dates_obj:
        default_date = max(available_dates_obj)  # Sélection automatique de la date la plus récente
        
        selected_date_obj = st.date_input(
            "Sélectionner la date",
            value=default_date,
            min_value=None,
            max_value=datetime.now().date(),
            format="DD/MM/YYYY"
        )
        
        selected_date_str = selected_date_obj.strftime('%d/%m/%Y')
        
        if selected_date_str not in sorted_dates_str:
            st.warning("La date sélectionnée n'est pas disponible dans le fichier. L'export sera vide si vous continuez.")
        
        download_file_name = f"{formatted_name}.xlsx" if formatted_name else "export_dyostem.xlsx"
        
        st.download_button(
            label="Générer et télécharger le fichier",
            data=process_file(file_content, selected_date_str),
            file_name=download_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Aucune date valide trouvée dans le fichier.")
else:
    if not displayed_options:
        st.info("Aucun fichier CSV Moûts trouvé dans le répertoire Export. Veuillez en charger un.")

# Enable auto-refresh every 5 seconds (5000 ms) to detect new files transparently
st_autorefresh(interval=5000, limit=None, key="directory_refresh")