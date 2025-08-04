import pandas as pd

def demarrage():
    print("ce logiciel fonctionne avec des fichier extension .xlsx\n\n")

    Nom_Source = input("Entrer le nom fichier source (sans extension)   ") + ".xlsx"
    import glob

    
    Nom_Cible = input("Entrer le nom fichier cible (sans extension)   ") + ".xlsx"

    
    return Nom_Cible, Nom_Source



from openpyxl import *
Nom_colonne = ["Nom de parcelle", "Cépage", "Date analyse", "Code échantillon", "Quantité Sucre (mg/baie)", "TAP (% vol)", "Acidité totale (g H2SO4/l)", "pH", "Acide malique (g/l)", "Acide tartrique", "Azote assimilable (mg/l)", "Potassium (g/l)", "Anthocyanes (mg/l)"]
def create_file(Nomcolonne):
    wb = Workbook()
    ws = wb.active

    import string
    alphabet = list(string.ascii_uppercase)[:len(Nom_colonne)]

    #print(len(alphabet))
    #print(len(Nom_colonne))

    for count, i in enumerate(Nom_colonne):
        ws[alphabet[count]+str(1)] = i
    return ws, wb


#importants
Nom_Cible, Nom_source = demarrage()
ws, wb = create_file(Nom_colonne)
#important





wb2 = load_workbook(filename= Nom_source)
ws2 = wb2[wb2.sheetnames[0]]


max_line = len(ws2["C"])
max_colm = len(ws2["1"])



ligne_ws2 = 1
from datetime import datetime

#voir les dates
dates = []
print("\n\n les dates disponibles sont:\n\n")
for i in range(max_line):
    if ws2["B"+str(ligne_ws2)].value == "Moûts":
        if datetime.strptime(str(ws2["D"+str(ligne_ws2)].value)[0:10],'%Y-%m-%d').strftime('%d/%m/%Y') not in dates:
            dates.append(datetime.strptime(str(ws2["D"+str(ligne_ws2)].value)[0:10],'%Y-%m-%d').strftime('%d/%m/%Y'))
            print(datetime.strptime(str(ws2["D"+str(ligne_ws2)].value)[0:10],'%Y-%m-%d').strftime('%d/%m/%Y'))
    ligne_ws2 += 1
date = input("entrer la date de moux  jj/mm/aaaa  ")

ligne_ws2 = 1
ligne_ws = 1
for i in range(max_line):
    if ws2["B"+str(ligne_ws2)].value == "Moûts" and datetime.strptime(str(ws2["D"+str(ligne_ws2)].value)[0:10],'%Y-%m-%d').strftime('%d/%m/%Y') == date:
        ligne_ws += 1
        ws["A"+str(ligne_ws)].value = ws2["C"+str(ligne_ws2)].value
        ws["B"+str(ligne_ws)].value = "Sauvignon blanc"
        ws["C"+str(ligne_ws)].value = datetime.strptime(str(ws2["D"+str(ligne_ws2)].value)[0:10],'%Y-%m-%d').strftime('%d/%m/%Y')
        ws["E"+str(ligne_ws)].value = ws2["E"+str(ligne_ws2)].value
        ws["F"+str(ligne_ws)].value = ws2["F"+str(ligne_ws2)].value
        ws["G"+str(ligne_ws)].value = ws2["G"+str(ligne_ws2)].value
        ws["H"+str(ligne_ws)].value = ws2["H"+str(ligne_ws2)].value
        ws["I"+str(ligne_ws)].value = ws2["I"+str(ligne_ws2)].value
        ws["J"+str(ligne_ws)].value = ws2["J"+str(ligne_ws2)].value
        ws["K"+str(ligne_ws)].value = ws2["K"+str(ligne_ws2)].value
        ws["L"+str(ligne_ws)].value = ws2["N"+str(ligne_ws2)].value
        ws["M"+str(ligne_ws)].value = ws2["P"+str(ligne_ws2)].value
    ligne_ws2 += 1

wb.save(Nom_Cible)

input('\n\nAppuyer sur entrée ...')
